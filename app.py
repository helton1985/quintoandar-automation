from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from werkzeug.utils import secure_filename
import os
import time
import threading
import json
from datetime import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'helton1985_21081985@_secret_key_production')

# Configurações
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# Criar diretório de uploads se não existir
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Variáveis globais para controle
automation_thread = None
automation_status = {
    'running': False,
    'progress': 0,
    'current_record': 0,
    'total_records': 0,
    'success_count': 0,
    'error_count': 0,
    'logs': [],
    'start_time': None,
    'duplicate_phones': []
}

def log_message(message, level='info'):
    """Adiciona mensagem aos logs com timestamp"""
    timestamp = datetime.now().strftime('%H:%M:%S')
    log_entry = f"[{timestamp}] {message}"
    automation_status['logs'].append(log_entry)
    print(log_entry)
    
    # Manter apenas últimos 100 logs
    if len(automation_status['logs']) > 100:
        automation_status['logs'] = automation_status['logs'][-100:]

def allowed_file(filename):
    """Verifica se arquivo tem extensão permitida"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_excel_data(file_path):
    """Processa arquivo Excel com mapeamento mais flexível"""
    try:
        log_message(f"📄 Carregando arquivo: {file_path}")
        
        # Carregar workbook
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # Ler cabeçalhos
        headers = []
        for cell in ws[1]:
            headers.append(cell.value if cell.value else '')
        
        log_message(f"📋 Colunas encontradas: {headers}")
        
        # Mapeamento de colunas com múltiplas possibilidades
        column_mapping = {
            'endereco': ['endereço', 'endereco', 'address', 'rua', 'logradouro'],
            'numero': ['número', 'numero', 'number', 'num', 'n'],
            'complemento': ['complemento', 'compl', 'complement'],
            'proprietario': ['proprietário', 'proprietario', 'nome', 'owner', 'name'],
            'telefone': ['telefone', 'celular', 'phone', 'fone', 'celular/telefone'],
            'email': ['email', 'e-mail', 'mail', 'correio']
        }
        
        # Mapear colunas
        mapped_columns = {}
        for key, possible_names in column_mapping.items():
            for i, header in enumerate(headers):
                if header and any(name.lower() in header.lower() for name in possible_names):
                    mapped_columns[key] = i
                    log_message(f"✅ Mapeado '{key}' → coluna '{header}' (índice {i})")
                    break
        
        log_message(f"🗂️ Mapeamento final: {mapped_columns}")
        
        # Processar dados
        records = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or all(cell is None or cell == '' for cell in row):
                continue
            
            # Extrair dados conforme mapeamento
            record = {}
            for field, col_idx in mapped_columns.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    record[field] = str(value).strip() if value else ''
            
            # Validar registro
            if record.get('proprietario') and record.get('endereco'):
                records.append(record)
                if len(records) == 1:
                    log_message(f"📝 Primeiro registro: {record}")
                if len(records) <= 3:
                    log_message(f"✅ Registro {len(records)} válido: {record.get('proprietario')}")
        
        log_message(f"📊 Processamento concluído: {len(records)} registros válidos de {row_num-1} total")
        return records
        
    except Exception as e:
        log_message(f"❌ Erro ao processar Excel: {str(e)}")
        return []

class QuintoAndarAutomation:
    def __init__(self):
        self.driver = None
        self.wait = None
        
    def setup_driver(self):
        """Configura o driver do Chrome para produção"""
        try:
            log_message("🔧 Configurando navegador Chrome...")
            
            # Configurações do Chrome
            chrome_options = Options()
            chrome_options.add_argument('--headless')  # Executar sem interface gráfica
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--disable-extensions')
            chrome_options.add_argument('--disable-plugins')
            chrome_options.add_argument('--disable-images')
            chrome_options.add_argument('--window-size=1920,1080')
            chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36')
            
            # Configurar service
            service = Service(ChromeDriverManager().install())
            
            # Criar driver
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.wait = WebDriverWait(self.driver, 15)
            
            log_message("✅ Chrome WebDriver configurado com sucesso!")
            return True
            
        except Exception as e:
            log_message(f"❌ Erro ao configurar Chrome: {str(e)}")
            return False
    
    def access_site(self):
        """Acessa o site do QuintoAndar"""
        try:
            log_message("🌐 Acessando site QuintoAndar...")
            self.driver.get("https://indicaai.quintoandar.com.br/")
            
            # Aguardar carregamento
            time.sleep(5)
            
            # Verificar se carregou
            if "QuintoAndar" in self.driver.title or "indica" in self.driver.title.lower():
                log_message("✅ Site acessado com sucesso!")
                return True
            else:
                log_message("❌ Erro: Site não carregou corretamente")
                return False
                
        except Exception as e:
            log_message(f"❌ Erro ao acessar site: {str(e)}")
            return False
    
    def fill_form(self, record):
        """Preenche formulário com dados do registro"""
        try:
            log_message(f"📝 Preenchendo formulário: {record.get('proprietario')}")
            
            # Aguardar formulário carregar
            time.sleep(3)
            
            # Preencher endereço
            if record.get('endereco'):
                try:
                    endereco_field = self.wait.until(
                        EC.presence_of_element_located((By.NAME, "endereco"))
                    )
                    endereco_field.clear()
                    endereco_field.send_keys(record['endereco'])
                    log_message(f"✅ Endereço preenchido: {record['endereco']}")
                    time.sleep(1)
                except:
                    log_message(f"⚠️ Campo endereço não encontrado")
            
            # Preencher número
            if record.get('numero'):
                try:
                    numero_field = self.driver.find_element(By.NAME, "numero")
                    numero_field.clear()
                    numero_field.send_keys(record['numero'])
                    log_message(f"✅ Número preenchido: {record['numero']}")
                    time.sleep(0.5)
                except:
                    log_message(f"⚠️ Campo número não encontrado")
            
            # Preencher complemento
            if record.get('complemento'):
                try:
                    complemento_field = self.driver.find_element(By.NAME, "complemento")
                    complemento_field.clear()
                    complemento_field.send_keys(record['complemento'])
                    log_message(f"✅ Complemento preenchido: {record['complemento']}")
                    time.sleep(0.5)
                except:
                    log_message(f"⚠️ Campo complemento não encontrado")
            
            # Preencher proprietário
            if record.get('proprietario'):
                try:
                    proprietario_field = self.driver.find_element(By.NAME, "proprietario")
                    proprietario_field.clear()
                    proprietario_field.send_keys(record['proprietario'])
                    log_message(f"✅ Proprietário preenchido: {record['proprietario']}")
                    time.sleep(0.5)
                except:
                    log_message(f"⚠️ Campo proprietário não encontrado")
            
            # Preencher telefone
            if record.get('telefone'):
                try:
                    telefone_field = self.driver.find_element(By.NAME, "telefone")
                    telefone_field.clear()
                    telefone_field.send_keys(record['telefone'])
                    log_message(f"✅ Telefone preenchido: {record['telefone']}")
                    time.sleep(0.5)
                except:
                    log_message(f"⚠️ Campo telefone não encontrado")
            
            # Preencher email
            if record.get('email'):
                try:
                    email_field = self.driver.find_element(By.NAME, "email")
                    email_field.clear()
                    email_field.send_keys(record['email'])
                    log_message(f"✅ Email preenchido: {record['email']}")
                    time.sleep(0.5)
                except:
                    log_message(f"⚠️ Campo email não encontrado")
            
            # Submeter formulário
            try:
                submit_button = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit'], input[type='submit']")
                submit_button.click()
                log_message("✅ Formulário enviado")
                
                # Aguardar resposta
                time.sleep(4)
                
                # Verificar se foi cadastrado com sucesso
                page_source = self.driver.page_source.lower()
                if "sucesso" in page_source or "cadastrado" in page_source or "obrigado" in page_source:
                    log_message(f"✅ Cadastro realizado com sucesso: {record.get('proprietario')}")
                    return True
                else:
                    log_message(f"⚠️ Possível erro no cadastro: {record.get('proprietario')}")
                    return False
                    
            except Exception as e:
                log_message(f"❌ Erro ao enviar formulário: {str(e)}")
                return False
                
        except TimeoutException:
            log_message(f"⏱️ Timeout no cadastro: {record.get('proprietario')}")
            return False
        except Exception as e:
            log_message(f"❌ Erro no cadastro: {str(e)}")
            return False
    
    def process_records(self, records):
        """Processa lista de registros"""
        global automation_status
        
        log_message(f"🚀 Iniciando processamento de {len(records)} registros")
        
        # Configurar driver
        if not self.setup_driver():
            log_message("❌ Erro ao configurar Chrome - interrompendo")
            automation_status['running'] = False
            return
        
        # Acessar site
        if not self.access_site():
            log_message("❌ Erro ao acessar site - interrompendo")
            automation_status['running'] = False
            return
        
        # Verificar telefones duplicados
        phones = [record.get('telefone', '') for record in records]
        phone_counts = {}
        for phone in phones:
            if phone:
                phone_counts[phone] = phone_counts.get(phone, 0) + 1
        
        duplicates = [phone for phone, count in phone_counts.items() if count > 1]
        if duplicates:
            log_message(f"⚠️ Encontrados {len(duplicates)} telefones duplicados")
            automation_status['duplicate_phones'] = duplicates
        
        # Processar registros
        success_count = 0
        error_count = 0
        
        for i, record in enumerate(records, 1):
            if not automation_status['running']:
                log_message("⏹️ Automação interrompida pelo usuário")
                break
            
            # Atualizar status
            automation_status['current_record'] = i
            automation_status['progress'] = (i / len(records)) * 100
            
            log_message(f"📋 Processando registro {i}/{len(records)}")
            
            # Verificar telefone duplicado
            if record.get('telefone') in duplicates:
                log_message(f"⚠️ Telefone duplicado detectado: {record.get('telefone')}")
            
            # Processar registro
            if self.fill_form(record):
                success_count += 1
                automation_status['success_count'] = success_count
            else:
                error_count += 1
                automation_status['error_count'] = error_count
            
            # Pausa entre registros
            time.sleep(3)
        
        # Finalizar
        log_message(f"🎉 Processamento concluído! Sucessos: {success_count}, Erros: {error_count}")
        automation_status['running'] = False
        
        # Fechar driver
        if self.driver:
            self.driver.quit()
            log_message("🔧 Navegador Chrome fechado")
    
    def stop_automation(self):
        """Para a automação"""
        global automation_status
        automation_status['running'] = False
        
        if self.driver:
            self.driver.quit()
            log_message("🔧 Navegador Chrome fechado")

# Rotas Flask
@app.route('/')
def index():
    """Página de login"""
    return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    """Processa login"""
    username = request.form.get('username')
    password = request.form.get('password')
    
    # Verificar credenciais
    if username == 'helton1985' and password == '21081985@':
        session['logged_in'] = True
        log_message(f"✅ Login realizado: {username}")
        return redirect(url_for('dashboard'))
    else:
        log_message(f"❌ Tentativa de login inválida: {username}")
        return render_template('index.html', error='Credenciais inválidas')

@app.route('/dashboard')
def dashboard():
    """Dashboard principal"""
    if not session.get('logged_in'):
        return redirect(url_for('index'))
    
    return render_template('dashboard.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Processa upload de arquivo"""
    global automation_thread, automation_status
    
    if not session.get('logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Tipo de arquivo não permitido'}), 400
    
    # Salvar arquivo
    filename = secure_filename(file.filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_{filename}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)
    
    log_message(f"📁 Arquivo salvo: {filename}")
    
    # Processar dados
    records = process_excel_data(file_path)
    
    if not records:
        return jsonify({'error': 'Nenhum dado válido encontrado no arquivo Excel'}), 400
    
    log_message(f"📊 {len(records)} registros encontrados para processamento")
    
    # Resetar status
    automation_status = {
        'running': True,
        'progress': 0,
        'current_record': 0,
        'total_records': len(records),
        'success_count': 0,
        'error_count': 0,
        'logs': automation_status['logs'],  # Manter logs existentes
        'start_time': datetime.now().isoformat(),
        'duplicate_phones': []
    }
    
    # Iniciar automação em thread separada
    automation = QuintoAndarAutomation()
    automation_thread = threading.Thread(target=automation.process_records, args=(records,))
    automation_thread.daemon = True
    automation_thread.start()
    
    return jsonify({
        'message': 'Automação iniciada com sucesso',
        'total_records': len(records)
    })

@app.route('/status')
def get_status():
    """Retorna status da automação"""
    if not session.get('logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    return jsonify(automation_status)

@app.route('/stop', methods=['POST'])
def stop_automation():
    """Para a automação"""
    global automation_status
    
    if not session.get('logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    automation_status['running'] = False
    log_message("🛑 Automação interrompida pelo usuário")
    
    return jsonify({'message': 'Automação interrompida'})

@app.route('/health')
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy'})

@app.route('/logout')
def logout():
    """Logout"""
    session.clear()
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
